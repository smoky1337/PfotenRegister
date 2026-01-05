import os
from collections import Counter
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook
from typing import Optional
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
)
from flask_login import login_required
from werkzeug.utils import secure_filename
from sqlalchemy import tuple_

from ...helpers import generate_unique_code, roles_required
from ...models import Animal, Guest, Representative, db

admin_io_bp = Blueprint('admin_io', __name__, url_prefix='/admin')

# Expected column names/types from the Excel template
GUEST_DTYPES = {
    "nummer": str,
    "vorname": str,
    "nachname": str,
    "adresse": str,
    "ort": str,
    "plz": str,
    "festnetz": str,
    "mobil": str,
    "email": str,
    "geschlecht": str,
    "status": str,
    "beduerftigkeit": str,
    "dokumente": str,
    "notizen": str,
    "vertreter_name": str,
    "vertreter_telefon": str,
    "vertreter_email": str,
    "vertreter_adresse": str,
}
ANIMAL_DTYPES = {
    "gast_nummer": str,
    "art": str,
    "rasse": str,
    "name": str,
    "geschlecht": str,
    "active": str,
    "farbe": str,
    "kastriert": str,
    "identifikation": str,
    "geburtsdatum": str,
    "gewicht_oder_groesse": str,
    "krankheiten": str,
    "unvertraeglichkeiten": str,
    "futter": str,
    "vollversorgung": str,
    "zuletzt_gesehen": str,
    "steuerbescheid_bis": str,
    "tierarzt": str,
    "futtermengeneintrag": str,
    "notizen": str,
}

DATE_FIELDS_GUESTS = [
    "geburtsdatum",
    "eintritt",
    "austritt",
    "beduerftig_bis",
    "erstellt_am",
    "aktualisiert_am",
]
STRING_FIELDS_GUESTS = list(GUEST_DTYPES.keys())

DATE_FIELDS_ANIMALS = [
    "geburtsdatum",
    "zuletzt_gesehen",
    "erstellt_am",
    "aktualisiert_am",
    "steuerbescheid_bis",
]
STRING_FIELDS_ANIMALS = list(ANIMAL_DTYPES.keys())

TRUE_VALUES = {"1", "ja", "true", "wahr", "y", "yes", "aktiv", "active", "x"}
FALSE_VALUES = {"0", "nein", "false", "falsch", "n", "no", "inaktiv", "inactive"}


def _safe_tmp_path(filepath: str):
    if not filepath:
        return None
    return os.path.join("tmp", os.path.basename(filepath))


def _normalize_date(val):
    if pd.isna(val) or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        converted = pd.to_datetime(val, dayfirst=True, errors="coerce")
        return converted.date() if not pd.isna(converted) else None
    except Exception:
        return None


def _normalize_string(val):
    return val.strip() if isinstance(val, str) and val.strip() != "" else None


def _normalize_dataframe(df, date_fields, string_fields):
    for field in date_fields:
        if field in df.columns:
            df[field] = df[field].apply(_normalize_date)
    for field in string_fields:
        if field in df.columns:
            df[field] = df[field].apply(_normalize_string)
    return df


def _coerce_text(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, str):
        stripped = val.strip()
        return stripped if stripped else None
    if isinstance(val, (int,)):
        return str(val)
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).strip()
    return str(val).strip() or None


def _iter_excel_records(
    safe_path: str,
    sheet_name: str,
    expected_columns: list,
    date_fields: list,
    string_fields: list,
    limit: Optional[int] = None,
):
    """
    Stream rows from an Excel sheet as dicts (minimizes memory vs pandas DataFrames).
    - Header row is expected in row 1.
    - Columns not present in the sheet are returned as None.
    """
    wb = load_workbook(filename=safe_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return
        header_map = {}
        for idx, name in enumerate(header):
            key = _coerce_text(name)
            if key:
                header_map[key.strip().lower()] = idx

        date_fields_set = {c.lower() for c in date_fields}
        string_fields_set = {c.lower() for c in string_fields}
        expected_lower = [c.lower() for c in expected_columns]
        relevant_indices = {
            header_map[col_lower]
            for col_lower in expected_lower
            if col_lower in header_map
        }

        emitted = 0
        excel_rownum = 1  # header row
        seen_data = False
        empty_streak = 0
        for row in rows:
            excel_rownum += 1
            if limit is not None and emitted >= limit:
                break

            # Guardrail: skip fully empty rows (Excel sheets often have max_row ~ 1,048,576)
            # and stop after a long empty tail once we've seen real data.
            is_empty = True
            for idx in relevant_indices:
                if idx < len(row):
                    cell = row[idx]
                    if cell is None:
                        continue
                    if isinstance(cell, str) and cell.strip() == "":
                        continue
                    if isinstance(cell, float) and pd.isna(cell):
                        continue
                    is_empty = False
                    break
            if is_empty:
                if seen_data:
                    empty_streak += 1
                    if empty_streak >= 200:
                        break
                continue

            seen_data = True
            empty_streak = 0
            record = {}
            for col, col_lower in zip(expected_columns, expected_lower):
                idx = header_map.get(col_lower)
                val = row[idx] if idx is not None and idx < len(row) else None
                if col_lower in date_fields_set:
                    record[col] = _normalize_date(val)
                elif col_lower in string_fields_set:
                    record[col] = _coerce_text(val)
                else:
                    record[col] = val
            record["_rownum"] = excel_rownum
            yield record
            emitted += 1
    finally:
        wb.close()


def _parse_bool(val, default=None):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    if isinstance(val, bool):
        return val
    text = str(val).strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return default


def _map_gender(value):
    if not value:
        return "Unbekannt"
    text = str(value).strip().lower()
    mapping = {
        "frau": "Frau",
        "weiblich": "Frau",
        "w": "Frau",
        "mann": "Mann",
        "maennlich": "Mann",
        "männlich": "Mann",
        "m": "Mann",
        "divers": "Divers",
        "d": "Divers",
        "unbekannt": "Unbekannt",
        "u": "Unbekannt",
    }
    return mapping.get(text, "Unbekannt")


def _map_animal_sex(value):
    if not value:
        return "Unbekannt"
    text = str(value).strip().lower()
    mapping = {
        "m": "M",
        "male": "M",
        "mann": "M",
        "maennlich": "M",
        "männlich": "M",
        "w": "F",
        "f": "F",
        "weiblich": "F",
        "female": "F",
        "unknown": "Unbekannt",
        "unbekannt": "Unbekannt",
    }
    return mapping.get(text, "Unbekannt")


def _map_choice(value, allowed, default=None):
    if not value:
        return default
    cleaned = str(value).strip()
    if cleaned in allowed:
        return cleaned
    lower_map = {str(v).lower(): v for v in allowed}
    mapped = lower_map.get(cleaned.lower())
    return mapped if mapped in allowed else default


def _map_yes_no_unknown(value, default="Unbekannt"):
    return _map_choice(value, {"Ja", "Nein", "Unbekannt"}, default=default)


def _map_species(value):
    return _map_choice(value, {"Hund", "Katze", "Vogel", "Nager", "Sonstige"})


def _map_food_type(value):
    return _map_choice(value, {"Misch", "Trocken", "Nass", "Barf"})

@admin_io_bp.route("/import", methods=["GET", "POST"])
@roles_required("admin")
@login_required
def import_data():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.endswith(".xlsx"):
            flash("Bitte eine gültige .xlsx Datei hochladen.", "danger")
            return redirect(request.url)

        filepath = os.path.join("tmp", secure_filename(file.filename))
        os.makedirs("tmp", exist_ok=True)
        file.save(filepath)

        return redirect(url_for("admin_io.preview_import", filepath=filepath))

    return render_template("admin/import.html", title="Datenimport")

@admin_io_bp.route("/import/confirm", methods=["GET", "POST"])
@roles_required("admin")
@login_required
def confirm_import():
    filepath = request.args.get("filepath")
    safe_path = _safe_tmp_path(filepath)

    if not safe_path or not os.path.exists(safe_path):
        flash("Datei nicht gefunden.", "danger")
        return redirect(url_for("admin_io.import_data"))

    today = date.today()
    guest_map = {}
    skipped_guests = 0
    skipped_animals = 0

    skipped_guest_samples = []
    skipped_animal_samples = []

    BATCH_SIZE = 250
    batch_counter = 0

    # Pre-scan guest sheet to enforce uniqueness constraints before importing anything.
    guest_numbers_counter = Counter()
    name_pairs = set()
    try:
        for guest in _iter_excel_records(
            safe_path,
            sheet_name="gaeste",
            expected_columns=list(GUEST_DTYPES.keys()) + [c for c in DATE_FIELDS_GUESTS if c not in GUEST_DTYPES],
            date_fields=DATE_FIELDS_GUESTS,
            string_fields=STRING_FIELDS_GUESTS,
            limit=None,
        ):
            number = guest.get("nummer")
            firstname = guest.get("vorname")
            lastname = guest.get("nachname")
            if number:
                guest_numbers_counter[str(number)] += 1
            if firstname and lastname:
                name_pairs.add((firstname, lastname))
    except Exception as exc:
        flash(f"Fehler beim Einlesen der Datei: {exc}", "danger")
        return redirect(url_for("admin_io.import_data"))

    duplicate_numbers = [n for n, c in guest_numbers_counter.items() if c > 1]
    if duplicate_numbers:
        flash(
            "Import abgebrochen: Gastnummern müssen eindeutig sein. Doppelte Nummern in der Datei: "
            + ", ".join(duplicate_numbers[:30])
            + (" …" if len(duplicate_numbers) > 30 else ""),
            "danger",
        )
        return redirect(url_for("admin_io.preview_import", filepath=safe_path))

    existing_numbers = set()
    if guest_numbers_counter:
        numbers = list(guest_numbers_counter.keys())
        CHUNK = 500
        for i in range(0, len(numbers), CHUNK):
            chunk = numbers[i:i + CHUNK]
            existing_numbers.update(
                row[0]
                for row in Guest.query.with_entities(Guest.number).filter(Guest.number.in_(chunk)).all()
            )
    if existing_numbers:
        flash(
            "Import abgebrochen: Gastnummern existieren bereits in der Datenbank: "
            + ", ".join(sorted(existing_numbers)[:30])
            + (" …" if len(existing_numbers) > 30 else ""),
            "danger",
        )
        return redirect(url_for("admin_io.preview_import", filepath=safe_path))

    existing_name_pairs = set()
    if name_pairs:
        pairs = list(name_pairs)
        CHUNK = 500
        for i in range(0, len(pairs), CHUNK):
            chunk = pairs[i:i + CHUNK]
            existing_name_pairs.update(
                tuple(row)
                for row in db.session.query(Guest.firstname, Guest.lastname)
                .filter(tuple_(Guest.firstname, Guest.lastname).in_(chunk))
                .all()
            )

    try:
        # Guests (and representatives)
        seen_numbers_in_import = set()
        for guest in _iter_excel_records(
            safe_path,
            sheet_name="gaeste",
            expected_columns=list(GUEST_DTYPES.keys()) + [c for c in DATE_FIELDS_GUESTS if c not in GUEST_DTYPES],
            date_fields=DATE_FIELDS_GUESTS,
            string_fields=STRING_FIELDS_GUESTS,
            limit=None,
        ):
            number = guest.get("nummer")
            firstname = guest.get("vorname")
            lastname = guest.get("nachname")
            if not number or not firstname or not lastname:
                skipped_guests += 1
                if len(skipped_guest_samples) < 20:
                    skipped_guest_samples.append(number or "(ohne Nummer)")
                continue
            number_str = str(number)
            if number_str in seen_numbers_in_import:
                skipped_guests += 1
                if len(skipped_guest_samples) < 20:
                    skipped_guest_samples.append(f"{number_str} (doppelt)")
                continue
            seen_numbers_in_import.add(number_str)
            if (firstname, lastname) in existing_name_pairs:
                skipped_guests += 1
                if len(skipped_guest_samples) < 20:
                    skipped_guest_samples.append(f"{number_str} ({firstname} {lastname} existiert)")
                continue

            guest_id = generate_unique_code()
            guest_map[number_str] = guest_id
            db.session.add(
                Guest(
                    id=guest_id,
                    number=number_str,
                    firstname=firstname,
                    lastname=lastname,
                    address=guest.get("adresse"),
                    city=guest.get("ort"),
                    zip=guest.get("plz"),
                    phone=guest.get("festnetz"),
                    mobile=guest.get("mobil"),
                    email=guest.get("email"),
                    birthdate=guest.get("geburtsdatum"),
                    gender=_map_gender(guest.get("geschlecht")),
                    member_since=guest.get("eintritt") or today,
                    member_until=guest.get("austritt"),
                    status=_parse_bool(guest.get("status"), default=True),
                    indigence=guest.get("beduerftigkeit"),
                    indigent_until=guest.get("beduerftig_bis"),
                    documents=guest.get("dokumente"),
                    notes=guest.get("notizen"),
                    created_on=guest.get("erstellt_am") or today,
                    updated_on=guest.get("aktualisiert_am") or today,
                )
            )
            if any(
                guest.get(key)
                for key in ("vertreter_name", "vertreter_telefon", "vertreter_email", "vertreter_adresse")
            ):
                db.session.add(
                    Representative(
                        guest_id=guest_id,
                        name=guest.get("vertreter_name"),
                        phone=guest.get("vertreter_telefon"),
                        email=guest.get("vertreter_email"),
                        address=guest.get("vertreter_adresse"),
                    )
                )

            batch_counter += 1
            if batch_counter % BATCH_SIZE == 0:
                db.session.flush()
                db.session.expunge_all()

        # Animals
        for animal in _iter_excel_records(
            safe_path,
            sheet_name="tiere",
            expected_columns=list(ANIMAL_DTYPES.keys()) + [c for c in DATE_FIELDS_ANIMALS if c not in ANIMAL_DTYPES],
            date_fields=DATE_FIELDS_ANIMALS,
            string_fields=STRING_FIELDS_ANIMALS,
            limit=None,
        ):
            guest_id = guest_map.get(str(animal.get("gast_nummer") or ""))
            if not guest_id:
                skipped_animals += 1
                if len(skipped_animal_samples) < 20:
                    skipped_animal_samples.append(animal.get("name") or "(Tier ohne Name)")
                continue

            db.session.add(
                Animal(
                    guest_id=guest_id,
                    species=_map_species(animal.get("art")),
                    breed=animal.get("rasse"),
                    name=animal.get("name"),
                    sex=_map_animal_sex(animal.get("geschlecht")),
                    color=animal.get("farbe"),
                    castrated=_map_yes_no_unknown(animal.get("kastriert")),
                    identification=animal.get("identifikation"),
                    birthdate=animal.get("geburtsdatum"),
                    weight_or_size=animal.get("gewicht_oder_groesse"),
                    illnesses=animal.get("krankheiten"),
                    allergies=animal.get("unvertraeglichkeiten"),
                    food_type=_map_food_type(animal.get("futter")),
                    complete_care=_map_yes_no_unknown(animal.get("vollversorgung")),
                    last_seen=animal.get("zuletzt_gesehen"),
                    veterinarian=animal.get("tierarzt"),
                    food_amount_note=animal.get("futtermengeneintrag"),
                    note=animal.get("notizen"),
                    status=_parse_bool(animal.get("active"), default=True),
                    tax_until=animal.get("steuerbescheid_bis"),
                    created_on=animal.get("erstellt_am") or today,
                    updated_on=animal.get("aktualisiert_am") or today,
                )
            )
            batch_counter += 1
            if batch_counter % BATCH_SIZE == 0:
                db.session.flush()
                db.session.expunge_all()

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f"Fehler beim Import: {exc}", "danger")
        return redirect(url_for("admin_io.preview_import", filepath=safe_path))

    if skipped_guests:
        flash(
            f"{skipped_guests} Gäste wurden übersprungen, da Pflichtfelder fehlen."
            + (f" Beispiele: {', '.join(skipped_guest_samples)}" if skipped_guest_samples else ""),
            "warning",
        )
    if skipped_animals:
        flash(
            f"{skipped_animals} Tiere wurden übersprungen, weil keine passende Gastnummer gefunden wurde."
            + (f" Beispiele: {', '.join(skipped_animal_samples)}" if skipped_animal_samples else ""),
            "warning",
        )

    flash("Import erfolgreich durchgeführt.", "success")
    return redirect(url_for("admin_io.import_data"))


@admin_io_bp.route("/import/preview")
@roles_required("admin")
@login_required
def preview_import():
    filepath = request.args.get("filepath")
    safe_path = _safe_tmp_path(filepath)

    if not filepath or not safe_path:
        flash(f"Kein Dateipfad angegeben. {filepath} {safe_path}", "danger")
        return redirect(url_for("admin_io.import_data"))

    if not os.path.exists(safe_path):
        flash("Die Datei wurde nicht gefunden.", "danger")
        return redirect(url_for("admin_io.import_data"))

    PREVIEW_LIMIT = 50

    # Validation: duplicates, existing entries, and missing references
    missing_guest_number = []
    missing_guest_reference = []

    guest_preview = []
    animal_preview = []

    guest_count = 0
    animal_count = 0

    guest_numbers_counter = Counter()
    guest_numbers_set = set()

    try:
        for guest in _iter_excel_records(
            safe_path,
            sheet_name="gaeste",
            expected_columns=list(GUEST_DTYPES.keys()) + [c for c in DATE_FIELDS_GUESTS if c not in GUEST_DTYPES],
            date_fields=DATE_FIELDS_GUESTS,
            string_fields=STRING_FIELDS_GUESTS,
            limit=None,
        ):
            guest_count += 1
            nummer = guest.get("nummer")
            if nummer:
                guest_numbers_counter[str(nummer)] += 1
                guest_numbers_set.add(str(nummer))
            if len(guest_preview) < PREVIEW_LIMIT:
                guest_preview.append(guest)

        for idx, row in enumerate(
            _iter_excel_records(
                safe_path,
                sheet_name="tiere",
                expected_columns=list(ANIMAL_DTYPES.keys()) + [c for c in DATE_FIELDS_ANIMALS if c not in ANIMAL_DTYPES],
                date_fields=DATE_FIELDS_ANIMALS,
                string_fields=STRING_FIELDS_ANIMALS,
                limit=None,
            )
        ):
            animal_count += 1
            guest_number = row.get("gast_nummer")
            rownum = row.get("_rownum") or (idx + 2)
            guest_number_text = str(guest_number).strip() if guest_number is not None and not pd.isna(guest_number) else ""
            if not guest_number_text:
                if len(missing_guest_number) < 20:
                    missing_guest_number.append(rownum)
            elif guest_number_text not in guest_numbers_set:
                if len(missing_guest_reference) < 20:
                    missing_guest_reference.append(f"{guest_number_text} (Zeile {rownum})")
            if len(animal_preview) < PREVIEW_LIMIT:
                animal_preview.append(row)
    except Exception as exc:
        flash(f"Fehler beim Einlesen der Datei: {exc}", "danger")
        return redirect(url_for("admin_io.import_data"))

    duplicates = [n for n, count in guest_numbers_counter.items() if count > 1]
    if duplicates:
        flash(
            f"Die folgenden Gastnummern sind mehrfach in der Datei vorhanden: {', '.join(duplicates[:30])}"
            + (" …" if len(duplicates) > 30 else ""),
            "danger",
        )

    existing_numbers = []
    if guest_numbers_set:
        numbers = sorted(guest_numbers_set)
        CHUNK = 500
        for i in range(0, len(numbers), CHUNK):
            chunk = numbers[i:i + CHUNK]
            existing_numbers.extend(
                [row[0] for row in Guest.query.with_entities(Guest.number).filter(Guest.number.in_(chunk)).all()]
            )
        if existing_numbers:
            flash(
                f"Die folgenden Gastnummern existieren bereits in der Datenbank: {', '.join(existing_numbers[:30])}"
                + (" …" if len(existing_numbers) > 30 else ""),
                "danger",
            )

    existing_guests = []
    for guest in guest_preview:
        if not guest.get("vorname") or not guest.get("nachname") or not guest.get("adresse"):
            continue
        exists = Guest.query.filter_by(
            firstname=guest["vorname"],
            lastname=guest["nachname"],
            address=guest["adresse"],
        ).first()
        if exists:
            existing_guests.append(f'{guest["vorname"]} {guest["nachname"]} ({guest["adresse"]})')

    if existing_guests:
        flash(
            f"Folgende Gäste existieren bereits (aus der Vorschau): {', '.join(existing_guests[:20])}"
            + (" …" if len(existing_guests) > 20 else ""),
            "warning",
        )

    if missing_guest_number:
        flash(
            f"Einige Tiere haben keine gültige Gastnummer (z. B. Zeilen: {', '.join(map(str, missing_guest_number))}"
            + (" …" if len(missing_guest_number) >= 20 else "")
            + ").",
            "danger",
        )
    if missing_guest_reference:
        flash(
            f"Tiere ohne passende Gastnummer in dieser Datei: {', '.join(missing_guest_reference)}"
            + (" …" if len(missing_guest_reference) >= 20 else "")
            + ". "
            f"Diese Tiere werden übersprungen.",
            "warning",
        )

    return render_template(
        "admin/import_preview.html",
        guests=guest_preview,
        animals=animal_preview,
        guest_count=guest_count,
        animal_count=animal_count,
        preview_limit=PREVIEW_LIMIT,
        filepath=safe_path,
        title="Importvorschau",
    )


@admin_io_bp.route("/export", methods=["GET", "POST"])
@roles_required("admin")
@login_required
def export_data():
    """Render export UI (GET) or generate an Excel export (POST) using SQLAlchemy.
    - Each selected table is exported to its own sheet.
    - Selected columns are respected; unknown columns are ignored safely.
    - Uses ORM models and `load_only` for efficient column selection.
    """
    from flask import send_file
    from io import BytesIO
    from sqlalchemy.orm import load_only
    from datetime import date
    from ...models import Guest, Animal, Payment, Message, Representative, db

    if request.method == "GET":
        return render_template("admin/export.html", title="Daten exportieren")

    # POST: collect selection from the form
    include_header = request.form.get("include_header") is not None
    selections = {
        "guests": request.form.getlist("fields[guests][]"),
        "animals": request.form.getlist("fields[animals][]"),
        "payments": request.form.getlist("fields[payments][]"),
        "messages": request.form.getlist("fields[messages][]"),
        "representatives": request.form.getlist("fields[representatives][]"),
    }
    # Map from UI field keys -> ORM attribute names (identity mapping, English names)
    FIELD_MAP = {
        "guests": {
            "address": "address",
            "birthdate": "birthdate",
            "city": "city",
            "created_on": "created_on",
            "documents": "documents",
            "email": "email",
            "firstname": "firstname",
            "gender": "gender",
            "guest_card_printed_on": "guest_card_printed_on",
            "guest_card_emailed_on": "guest_card_emailed_on",
            "id": "id",
            "indigence": "indigence",
            "indigent_until": "indigent_until",
            "lastname": "lastname",
            "member_since": "member_since",
            "member_until": "member_until",
            "mobile": "mobile",
            "notes": "notes",
            "number": "number",
            "phone": "phone",
            "status": "status",
            "updated_on": "updated_on",
            "zip": "zip",
        },
        "animals": {
            "allergies": "allergies",
            "birthdate": "birthdate",
            "breed": "breed",
            "castrated": "castrated",
            "color": "color",
            "complete_care": "complete_care",
            "created_on": "created_on",
            "died_on": "died_on",
            "food_amount_note": "food_amount_note",
            "food_type": "food_type",
            "guest_id": "guest_id",
            "id": "id",
            "identification": "identification",
            "illnesses": "illnesses",
            "last_seen": "last_seen",
            "name": "name",
            "note": "note",
            "pet_registry": "pet_registry",
            "sex": "sex",
            "species": "species",
            "status": "status",
            "tax_until": "tax_until",
            "updated_on": "updated_on",
            "veterinarian": "veterinarian",
            "weight_or_size": "weight_or_size",
        },
        "payments": {
            "comment": "comment",
            "created_on": "created_on",
            "food_amount": "food_amount",
            "guest_id": "guest_id",
            "id": "id",
            "other_amount": "other_amount",
            "paid": "paid",
            "paid_on": "paid_on",
        },
        "messages": {
            "completed": "completed",
            "content": "content",
            "created_by": "created_by",
            "created_on": "created_on",
            "guest_id": "guest_id",
            "id": "id",
        },
        "representatives": {
            "address": "address",
            "email": "email",
            "guest_id": "guest_id",
            "id": "id",
            "name": "name",
            "phone": "phone",
        },
    }
    TABLE_MODEL = {
        "guests": (Guest, "guests"),
        "animals": (Animal, "animals"),
        "payments": (Payment, "payments"),
        "messages": (Message, "messages"),
        "representatives": (Representative, "representatives"),
    }

    # Helper: build DataFrame from ORM query
    def rows_to_df(model, cols_ui, table_key):
        if model is None or not cols_ui:
            return None
        ui_to_attr = FIELD_MAP.get(table_key, {})
        attr_names = [ui_to_attr.get(c) for c in cols_ui]
        attr_names = [a for a in attr_names if a and hasattr(model, a)]
        if not attr_names:
            return None
        load_columns = [getattr(model, a) for a in attr_names]
        q = db.session.query(model).options(load_only(*load_columns))
        results = q.all()
        rows = []
        for obj in results:
            row = []
            for ui_col in cols_ui:
                attr = ui_to_attr.get(ui_col)
                val = getattr(obj, attr, None) if attr and hasattr(obj, attr) else None
                if hasattr(val, "isoformat"):
                    try:
                        val = val.isoformat()
                    except Exception:
                        pass
                row.append(val)
            rows.append(row)
        headers = cols_ui if include_header else None
        df = pd.DataFrame(rows, columns=headers if headers else None)
        return df

    # Build the Excel in-memory
    output = BytesIO()
    any_sheet = False
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for key, (model, sheet_name) in TABLE_MODEL.items():
            cols = selections.get(key) or []
            df = rows_to_df(model, cols, key)
            if df is None:
                continue
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=include_header)
            any_sheet = True
    if not any_sheet:
        flash("Bitte wähle mindestens eine Spalte aus.", "warning")
        return render_template("admin/export.html", title="Daten exportieren")

    output.seek(0)
    filename = f"pfotenregister_export_{date.today().isoformat()}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)

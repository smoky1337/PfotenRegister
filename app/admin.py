import io
import os
from datetime import date

import pandas as pd
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    current_app,
)
from werkzeug.security import generate_password_hash
from flask_login import current_user, login_required
from werkzeug.utils import secure_filename
from collections import defaultdict

from .auth import get_user_by_username
from .models import db, Guest, Animal, User, FoodHistory, Payments, FieldRegistry, Setting

from .helpers import roles_required, get_form_value

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")






@admin_bp.route("/")
@login_required
@roles_required("admin")
def dashboard():
    from datetime import date, timedelta

    total_guests = Guest.query.count()
    active_guests = Guest.query.filter_by(status="Aktiv").count()

    last_30_days = date.today() - timedelta(days=30)
    recent_guests = (
        FoodHistory.query.filter(FoodHistory.distributed_on >= last_30_days)
        .with_entities(FoodHistory.guest_id)
        .distinct()
        .count()
    )

    total_animals = Animal.query.count()
    animals_by_type = (
        Animal.query.with_entities(Animal.species, db.func.count().label("count"))
        .group_by(Animal.species)
        .all()
    )

    top_guests_by_visits = (
        db.session.query(Guest.number, Guest.firstname, Guest.lastname, db.func.count(FoodHistory.id).label("besuche"))
        .outerjoin(FoodHistory, Guest.id == FoodHistory.guest_id)
        .group_by(Guest.id)
        .order_by(db.desc("besuche"))
        .limit(10)
        .all()
    )

    payment_trends = (
        db.session.query(
            FoodHistory.distributed_on,
            db.func.coalesce(db.func.sum(Payments.food_amount), 0).label("Futtersumme"),
            db.func.coalesce(db.func.sum(Payments.other_amount), 0).label("Andere"),
        )
        .outerjoin(Payments, (FoodHistory.guest_id == Payments.guest_id) & (Payments.paid_on == FoodHistory.distributed_on))
        .group_by(FoodHistory.distributed_on)
        .order_by(FoodHistory.distributed_on.asc())
        .limit(30)
        .all()
    )

    return render_template(
        "admin/dashboard.html",
        current_user=current_user,
        total_guests=total_guests,
        active_guests=active_guests,
        recent_guests=recent_guests,
        total_animals=total_animals,
        animals_by_type=animals_by_type,
        top_guests_by_visits=top_guests_by_visits,
        payment_trends=payment_trends,
        title="Dashboard"
    )

@admin_bp.route("/list_users")
@roles_required("admin")
@login_required
def list_users():
    users = User.query.all()
    return render_template(
        "admin/list_users.html", users=users, title="Benutzerverwaltung"
    )


@admin_bp.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@roles_required("admin")
@login_required
def edit_user(user_id):
    user = User.query.get(user_id)
    if not user:
        flash("Benutzer nicht gefunden.", "danger")
        return redirect(url_for("admin.list_users"))
    if request.method == "POST":
        username = get_form_value("username")
        role = get_form_value("role")
        new_password = get_form_value("password")
        realname = get_form_value("realname")
        user.username = username
        user.role = role
        user.realname = realname
        if new_password:
            user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        flash("Benutzer erfolgreich aktualisiert.", "success")
        return redirect(url_for("admin.list_users"))
    else:
        return render_template(
            "admin/edit_user.html", user=user, title="Benutzer bearbeiten"
        )


@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
@roles_required("admin")
@login_required
def delete_user(user_id):
    User.query.filter_by(id=user_id).delete()
    db.session.commit()
    flash("Benutzer erfolgreich gelöscht.", "success")
    return redirect(url_for("admin.list_users"))


@admin_bp.route("/field_visibility", methods=["POST"])
@roles_required("admin")
@login_required
def update_field_visibility():
    visible_ids = set(map(int, request.form.getlist("visible_fields")))
    all_fields = FieldRegistry.query.all()
    for field in all_fields:
        field.globally_visible = field.id in visible_ids
        new_level = get_form_value(f"visibility_level_{field.id}")
        if new_level != field.visibility_level:
            field.visibility_level = new_level
        new_ui = get_form_value(f"ui_label_{field.id}")
        if new_ui != field.ui_label:
            field.ui_label = new_ui
    db.session.commit()
    flash("Feldsichtbarkeit wurde aktualisiert.", "success")
    return redirect(url_for("admin.edit_settings"))

@admin_bp.route("/settings", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def edit_settings():
    if request.method == "POST":
        # Gehe alle Settings durch und update sie
        for key in request.form:
            value = get_form_value(key)
            Setting.query.filter_by(key=key).update(value)


        current_app.refresh_settings()

        flash("Einstellungen wurden gespeichert und aktualisiert.", "success")
        return redirect(url_for("admin.edit_settings"))
    else:
        # GET: zeige aktuelle Settings
        # group optional fields by model
        field_registry = defaultdict(list)
        query = FieldRegistry.query.order_by(FieldRegistry.model_name,
                                                                           FieldRegistry.field_name).all()
        for field in query:
            field_registry[field.model_name].append(field)

        settings = Setting.query.all()
        settings = {item["setting_key"]: item for item in settings}
        return render_template("admin/edit_settings.html", settings=settings, field_registry=field_registry)


@admin_bp.route("/users/register", methods=["GET", "POST"])
@roles_required("admin")
@login_required
def register_user():
    if request.method == "POST":
        username = get_form_value("username")
        password = get_form_value("password")
        role = get_form_value("role")
        realname = get_form_value("realname")
        if not username or not password or not role:
            flash("Bitte alle Felder ausfüllen.", "danger")
            return redirect(url_for("auth.create_user"))
        if get_user_by_username(username):
            flash("Benutzername existiert bereits.", "danger")
            return redirect(url_for("auth.create_user"))
        password_hash = generate_password_hash(password)
        user = User(username=username, password_hash=password_hash, role=role, realname=realname)
        db.session.add(user)
        db.session.commit()
        flash("Benutzer erfolgreich angelegt.", "success")
        return redirect(url_for("admin.list_users"))
    return render_template("admin/register_user.html", title="Benutzer anlegen")




@admin_bp.route("/import", methods=["GET", "POST"])
@roles_required("admin")
@login_required
def import_data():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.endswith(".xlsx"):
            flash("Bitte eine gültige .xlsx Datei hochladen.", "danger")
            return redirect(request.url)

        filepath = os.path.join("tmp", secure_filename(file.filename))
        os.makedirs("tmp", exist_ok=True)
        file.save(filepath)

        return redirect(url_for("admin.preview_import", filepath=filepath))

    return render_template("admin/import.html", title="Datenimport")

@admin_bp.route("/import/confirm", methods=["GET", "POST"])
@roles_required("admin")
@login_required
def confirm_import():
    from .helpers import generate_unique_code
    from .db import db_cursor
    import pandas as pd
    import os

    filepath = request.args.get("filepath")
    safe_path = os.path.join("tmp", os.path.basename(filepath))

    if not os.path.exists(safe_path):
        flash("Datei nicht gefunden.", "danger")
        return redirect(url_for("admin.import_data"))

    dtype_guests = {
        "nummer": str,
        "vorname": str,
        "nachname": str,
        "adresse": str,
        "ort": str,
        "plz": str,
        "festnetz": str,
        "mobil": str,
        "email": str,
        "geschlecht": str,
        "status": str,
        "beduerftigkeit": str,
        "dokumente": str,
        "notizen": str,
        "vertreter_name": str,
        "vertreter_telefon": str,
        "vertreter_email": str,
        "vertreter_adresse": str
    }
    dtype_animals = {
        "gast_nummer": str,
        "art": str,
        "rasse": str,
        "name": str,
        "geschlecht": str,
        "farbe": str,
        "identifikation": str,
        "kastriert": str,
        "futter": str,
        "vollversorgung": str,
        "notizen": str,
        "active": str
    }
    df_guests = pd.read_excel(safe_path, sheet_name="gaeste", dtype=dtype_guests)
    df_animals = pd.read_excel(safe_path, sheet_name="tiere", dtype=dtype_animals)

    def normalize_date(val):
        if pd.isna(val) or val == "":
            return None
        try:
            return pd.to_datetime(val, dayfirst=True).date()
        except Exception:
            return None

    def normalize_string(val):
        return val.strip() if isinstance(val, str) and val.strip() != "" else None

    for col in df_guests.columns:
        if col in ["geburtsdatum", "eintritt", "austritt", "beduerftig_bis", "erstellt_am", "aktualisiert_am"]:
            df_guests[col] = df_guests[col].apply(normalize_date)
        else:
            df_guests[col] = df_guests[col].apply(normalize_string)

    for col in df_animals.columns:
        if col in ["geburtsdatum", "zuletzt_gesehen", "erstellt_am", "aktualisiert_am", "steuerbescheid_bis"]:
            df_animals[col] = df_animals[col].apply(normalize_date)
        else:
            df_animals[col] = df_animals[col].apply(normalize_string)

    guest_map = {}

    with db_cursor() as cursor:
        today = date.today()
        for guest in df_guests.to_dict(orient="records"):
            print(guest)
            guest_id = generate_unique_code()
            status = guest.get("status")
            if not status:
                status = "Aktiv"
            guest_map[guest.get("nummer")] = guest_id
            cursor.execute("""
                INSERT INTO gaeste (
                    id, nummer, vorname, nachname, adresse, ort, plz,
                    festnetz, mobil, email, geburtsdatum, geschlecht,
                    eintritt, austritt, vertreter_name, vertreter_telefon,
                    vertreter_email, vertreter_adresse, status,
                    beduerftigkeit, beduerftig_bis, dokumente, notizen,
                    erstellt_am, aktualisiert_am
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s
                )
            """, (
                guest_id,
                guest.get("nummer"), guest.get("vorname"), guest.get("nachname"), guest.get("adresse"), guest.get("ort"), guest.get("plz"),
                guest.get("festnetz"), guest.get("mobil"), guest.get("email"), guest.get("geburtsdatum"), guest.get("geschlecht"),
                guest.get("eintritt"), guest.get("austritt"), guest.get("vertreter_name"), guest.get("vertreter_telefon"),
                guest.get("vertreter_email"), guest.get("vertreter_adresse"), status,
                guest.get("beduerftigkeit"), guest.get("beduerftig"), guest.get("dokumente"), guest.get("notizen"),
                guest.get("erstellt_am") or today, guest.get("aktualisiert_am") or today
            ))

        for animal in df_animals.to_dict(orient="records"):
            guest_id = guest_map.get(animal.get("gast_nummer"))
            if not guest_id:
                continue  # Keine gültige Zuordnung gefunden

            cursor.execute("""
                INSERT INTO tiere (
                    guest_id, art, rasse, name, geschlecht, farbe, kastriert, identifikation,
                    geburtsdatum, gewicht_oder_groesse, krankheiten, unvertraeglichkeiten,
                    futter, vollversorgung, zuletzt_gesehen, tierarzt, futtermengeneintrag,
                    notizen, active, steuerbescheid_bis, erstellt_am, aktualisiert_am
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s
                )
            """, (
                guest_id, animal.get("art"), animal.get("rasse"), animal.get("name"), animal.get("geschlecht"), animal.get("farbe"),
                animal.get("kastriert"), animal.get("identifikation"), animal.get("geburtsdatum"),
                animal.get("gewicht_oder_groesse"), animal.get("krankheiten"), animal.get("unvertraeglichkeiten"),
                animal.get("futter"), animal.get("vollversorgung"), animal.get("zuletzt_gesehen"), animal.get("tierarzt"),
                animal.get("futtermengeneintrag"), animal.get("notizen"), animal.get("aktiv"), animal.get("steuerbescheid_bis"),
                animal.get("erstellt_am") or today, animal.get("aktualisiert_am") or today
            ))

    flash("Import erfolgreich durchgeführt.", "success")
    return redirect(url_for("admin.import_data"))


@admin_bp.route("/import/preview")
@roles_required("admin")
@login_required
def preview_import():
    filepath = request.args.get("filepath")

    if not filepath:
        flash("Kein Dateipfad angegeben.", "danger")
        return redirect(url_for("admin.import_data"))

    # Sicherheitshalber: Nur Zugriff auf tmp-Ordner zulassen
    safe_path = os.path.join("tmp", os.path.basename(filepath))
    if not os.path.exists(safe_path):
        flash("Die Datei wurde nicht gefunden.", "danger")
        return redirect(url_for("admin.import_data"))

    try:
        dtype_guests = {
            "nummer": str,
            "vorname": str,
            "nachname": str,
            "adresse": str,
            "ort": str,
            "plz": str,
            "festnetz": str,
            "mobil": str,
            "email": str,
            "geschlecht": str,
            "status": str,
            "dokumente": str,
            "notizen": str,
            "vertreter_name": str,
            "vertreter_telefon": str,
            "vertreter_email": str,
            "vertreter_adresse": str
        }
        dtype_animals = {
            "gast_nummer": str,
            "art": str,
            "rasse": str,
            "name": str,
            "geschlecht": str,
            "farbe": str,
            "identifikation": str,
            "kastriert": str,
            "futter": str,
            "vollversorgung": str,
            "notizen": str,
            "aktiv": str,
        }
        df_guests = pd.read_excel(safe_path, sheet_name="gaeste", dtype=dtype_guests)
        df_animals = pd.read_excel(safe_path, sheet_name="tiere", dtype=dtype_animals)
        print(df_guests)
        print(df_animals)
    except Exception as e:
        flash(f"Fehler beim Einlesen der Datei: {e}", "danger")
        return redirect(url_for("admin.import_data"))

    # --- Datenvalidierung und Normalisierung ---
    def normalize_date(val):
        if pd.isna(val) or val == "":
            return None
        try:
            return pd.to_datetime(val, dayfirst=True).date()
        except Exception:
            return None

    def normalize_string(val):
        return val.strip() if isinstance(val, str) and val.strip() != "" else None

    date_fields_guests = ["geburtsdatum", "eintritt", "austritt", "beduerftig_bis", "erstellt_am", "aktualisiert_am"]
    string_fields_guests = [
        "nummer",
        "vorname",
        "nachname",
        "adresse",
        "ort",
        "plz",
        "festnetz",
        "mobil",
        "email",
        "geschlecht",
        "status",
        "dokumente",
        "notizen",
        "vertreter_name",
        "vertreter_telefon",
        "vertreter_email",
        "vertreter_adresse",
    ]

    for field in date_fields_guests:
        if field in df_guests.columns:
            df_guests[field] = df_guests[field].apply(normalize_date)

    for field in string_fields_guests:
        if field in df_guests.columns:
            df_guests[field] = df_guests[field].apply(normalize_string)

    date_fields_animals = ["geburtsdatum", "zuletzt_gesehen", "erstellt_am", "aktualisiert_am", "steuerbescheid_bis"]
    string_fields_animals = [
        "gast_nummer",
        "art",
        "rasse",
        "name",
        "geschlecht",
        "farbe",
        "identifikation",
        "kastriert",
        "futter",
        "vollversorgung",
        "notizen",
    ]

    for field in date_fields_animals:
        if field in df_animals.columns:
            df_animals[field] = df_animals[field].apply(normalize_date)

    for field in string_fields_animals:
        if field in df_animals.columns:
            df_animals[field] = df_animals[field].apply(normalize_string)


    # Validation: Check for missing or empty gast_nummer in animals
    missing_guest_number = []
    for idx, row in df_animals.iterrows():
        if pd.isna(row.get("gast_nummer")) or str(row.get("gast_nummer")).strip() == "":
            missing_guest_number.append(idx + 2)  # Excel-style row number

    if missing_guest_number:
        flash(f"Einige Tiere haben keine gültige Gastnummer (z. B. Zeile(n): {', '.join(map(str, missing_guest_number))}).", "danger")


    # Check for duplicate guests by vorname, nachname, adresse
    with db_cursor() as cursor:
        # Check for duplicate guest numbers in the imported file
        duplicate_numbers = df_guests["nummer"].value_counts()
        duplicates = duplicate_numbers[duplicate_numbers > 1]
        if not duplicates.empty:
            flash(f"Die folgenden Gastnummern sind mehrfach in der Datei vorhanden: {', '.join(duplicates.index)}",
                  "danger")

        # Check for guest numbers that already exist in the database
        existing_numbers = []
        for nummer in df_guests["nummer"].dropna().unique():
            nummer = str(nummer)
            cursor.execute("SELECT id FROM gaeste WHERE nummer = %s", (nummer,))
            if cursor.fetchone():
                existing_numbers.append(nummer)

        if existing_numbers:
            flash(f"Die folgenden Gastnummern existieren bereits in der Datenbank: {', '.join(existing_numbers)}",
                  "danger")

        existing_guests = []
        for guest in df_guests.to_dict(orient="records"):
            if not guest.get("vorname") or not guest.get("nachname") or not guest.get("adresse"):
                continue  # skip if necessary fields are missing
            cursor.execute(
                "SELECT id FROM gaeste WHERE vorname = %s AND nachname = %s AND adresse = %s",
                (guest["vorname"], guest["nachname"], guest["adresse"])
            )
            if cursor.fetchone():
                existing_guests.append(f'{guest["vorname"]} {guest["nachname"]} ({guest["adresse"]})')

        if existing_guests:
            flash(f"Folgende Gäste existieren bereits: {', '.join(existing_guests)}", "warning")


    return render_template(
        "admin/import_preview.html",
        guests=df_guests.to_dict(orient="records"),
        animals=df_animals.to_dict(orient="records"),
        filepath=safe_path,
        title="Importvorschau"
    )

@admin_bp.route("/export", methods=["GET"])
@roles_required("admin")
@login_required
def export_data():
    from flask import send_file
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from io import BytesIO
    from datetime import date
    import pandas as pd

    template_path = os.path.join(current_app.root_path, "static", "output_template.xlsx")
    wb = load_workbook(template_path)

    ws_guests = wb["gaeste"]
    ws_animals = wb["tiere"]

    # Clear existing rows except for headers
    ws_guests.delete_rows(2, ws_guests.max_row)
    ws_animals.delete_rows(2, ws_animals.max_row)

    with db_cursor() as cursor:
        cursor.execute("SELECT * FROM gaeste")
        guests = cursor.fetchall()
        cursor.execute("SELECT * FROM tiere")
        animals = cursor.fetchall()

    df_guests = pd.DataFrame(guests)[[
        "nummer", "vorname", "nachname", "adresse", "plz", "ort", "festnetz", "mobil", "email",
        "geburtsdatum", "geschlecht", "eintritt", "austritt",
        "vertreter_name", "vertreter_telefon", "vertreter_email", "vertreter_adresse",
        "status", "beduerftigkeit", "beduerftig_bis", "dokumente", "notizen"
    ]]

    df_animals = pd.DataFrame(animals)

    # Merge animals with guest nummer
    guest_id_to_nummer = pd.DataFrame(guests)[["id", "nummer"]].rename(columns={"id": "guest_id"})
    df_animals = df_animals.merge(guest_id_to_nummer, on="guest_id", how="left")

    # test if theres any animals without guests
    if df_animals["nummer"].isna().any():
        print("[WARNUNG] Einige Tiere konnten keiner gültigen Gastnummer zugeordnet werden.")

    df_animals.drop(columns=["guest_id"], inplace=True)
    df_animals = df_animals[[
        "nummer", "art", "rasse", "name", "geschlecht", "farbe", "kastriert", "identifikation",
        "geburtsdatum", "gewicht_oder_groesse", "krankheiten", "unvertraeglichkeiten",
        "futter", "vollversorgung", "zuletzt_gesehen", "tierarzt", "futtermengeneintrag", "notizen"
    ]]
    df_animals.rename(columns={"nummer": "gast_nummer"}, inplace=True)

    for row in dataframe_to_rows(df_guests, index=False, header=False):
        ws_guests.append(row)

    for row in dataframe_to_rows(df_animals, index=False, header=False):
        ws_animals.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"pfotenregister_export_{date.today().isoformat()}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)


@admin_bp.route("/export_transactions", methods=["GET"])
@roles_required("admin")
@login_required
def export_transactions():
    from flask import send_file
    from io import BytesIO
    from datetime import datetime
    import pandas as pd

    from_date = request.args.get("from")
    to_date = request.args.get("to")

    if not from_date or not to_date:
        flash("Bitte gib ein gültiges Start- und Enddatum an.", "danger")
        return redirect(url_for("admin.dashboard"))

    try:
        from_dt = datetime.strptime(from_date, "%Y-%m-%d").date()
        to_dt = datetime.strptime(to_date, "%Y-%m-%d").date()
    except ValueError:
        flash("Ungültiges Datumsformat.", "danger")
        return redirect(url_for("admin.dashboard"))

    with db_cursor() as cursor:
        cursor.execute("""
            SELECT z.zahlungstag, g.nummer AS gast_nummer, g.vorname, g.nachname,
                   z.futter_betrag, z.zubehoer_betrag, z.kommentar
            FROM zahlungshistorie z
            JOIN gaeste g ON g.id = z.guest_id
            WHERE z.zahlungstag BETWEEN %s AND %s
            ORDER BY z.zahlungstag ASC
        """, (from_dt, to_dt))
        records = cursor.fetchall()

    df = pd.DataFrame(records)
    if not df.empty:
        total_futter = df["futter_betrag"].sum()
        total_zubehoer = df["zubehoer_betrag"].sum()
    else:
        # Leere DataFrame mit Spaltennamen
        df = pd.DataFrame(columns=[
            "zahlungstag", "gast_nummer", "vorname", "nachname",
            "futter_betrag", "zubehoer_betrag", "kommentar"
        ])
        total_futter = 0
        total_zubehoer = 0

    return render_template(
        "reports/transaction_report.html",
        transactions=records,
        current_user=current_user,
        from_date=from_dt.strftime('%d.%m.%Y'),
        to_date=to_dt.strftime('%d.%m.%Y'),
        total_futter=total_futter,
        total_zubehoer=total_zubehoer,
        now = datetime.today(),
        title="Zahlungsbericht"
    )

    filename = f"zahlungsexport_{from_dt.isoformat()}_bis_{to_dt.isoformat()}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)